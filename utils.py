from openpyxl import load_workbook

def read_student_list(class_name, subject):
    path = f'data\\{class_name}\\{subject}.xlsx'
    try:
        wb = load_workbook(path)
        sheet = wb.active
        students = []
        for row in range(2, sheet.max_row + 1):
            student_id = sheet.cell(row=row, column=1).value
            student_name = sheet.cell(row=row, column=2).value
            students.append({'id': student_id, 'name': student_name})
        return students
    except Exception as e:
        raise Exception(f"读取学生名单失败: {str(e)}")

def update_student_status(class_name, subject, student_id, assignment_name, status):
    path = f'data/{class_name}/{subject}.xlsx'
    try:
        wb = load_workbook(path)
        sheet = wb.active
        # 找到作业列
        col = None
        for idx in range(4, sheet.max_column + 1):
            if sheet.cell(row=1, column=idx).value == assignment_name:
                col = idx
                break
        if not col:
            raise Exception("作业不存在")
        # 更新状态
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == student_id:
                sheet.cell(row=row, column=col).value = status
                wb.save(path)
                return
        raise Exception("学生不存在")
    except Exception as e:
        raise Exception(f"更新状态失败: {str(e)}")

def read_assignment_list(class_name, subject):
    path = f'data/{class_name}/{subject}.xlsx'
    try:
        wb = load_workbook(path)
        sheet = wb.active
        assignments = []
        for col in range(4, sheet.max_column + 1):
            assignment = sheet.cell(row=1, column=col).value
            if assignment:
                assignments.append(assignment)
        return assignments
    except Exception as e:
        raise Exception(f"读取作业列表失败: {str(e)}")

def create_assignment(class_name, subject, assignment_name):
    path = f'data/{class_name}/{subject}.xlsx'
    try:
        wb = load_workbook(path)
        sheet = wb.active
        new_col = sheet.max_column + 1
        sheet.cell(row=1, column=new_col).value = assignment_name
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=new_col).value = ""
        wb.save(path)
        return True
    except Exception as e:
        raise Exception(f"创建作业失败: {str(e)}")
