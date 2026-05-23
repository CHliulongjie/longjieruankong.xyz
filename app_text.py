from flask import Flask, request, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
import sqlite3
import os
import logging
from waitress import serve
import openpyxl  # 添加openpyxl导入
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from flask_cors import CORS
from flask import render_template_string

app = Flask(__name__)
CORS(app)

# ===== 路径配置 =====
base_dir = os.path.dirname(os.path.abspath(__file__))
data_dir = os.path.join(base_dir, 'data')
os.makedirs(data_dir, exist_ok=True)

# ===== 数据库配置 =====
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(data_dir, "users.db").replace("\\", "/")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'data'
ALLOWED_EXTENSIONS = {'xlsx'}
db = SQLAlchemy(app)

class User(db.Model):
    __tablename__ = 'users'  # 显式指定表名
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    roles = db.Column(db.String(200))      # 新增字段
    permissions = db.Column(db.String(200)) # 新增字段

# ===== 工具函数 =====
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn

def get_file_path(class_name, subject):
    return os.path.join(app.config['UPLOAD_FOLDER'], class_name, f"{subject}.xlsx")
# ------------------- 核心接口 -------------------
@app.route('/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    user = User.query.filter_by(username=username, password=password).first()
    if user:
        return jsonify({
            'username': user.username,
            'message': '登录成功',
            'roles': user.roles.split(','),
            'permissions': user.permissions.split(',') if user.permissions else []
        }), 200
    else:
        return jsonify({'message': '用户名或密码错误'}), 401

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    roles = data.get('roles', 'student')  # 默认角色为学生
    
    if not username or not password:
        return jsonify({'message': '用户名和密码不能为空'}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({'message': '用户名已存在'}), 400

    new_user = User(
        username=username,
        password=password,
        roles=roles,
        permissions=''
    )
    db.session.add(new_user)
    db.session.commit()
    return jsonify({'message': '用户注册成功'}), 201

@app.route('/set_permissions', methods=['POST'])
def set_permissions():
    data = request.json
    username = data.get('username')
    roles = data.get('roles')
    permissions = data.get('permissions')

    user = User.query.filter_by(username=username).first()
    if not user:
        return jsonify({'message': '用户不存在'}), 404

    try:
        user.roles = roles
        user.permissions = permissions
        db.session.commit()
        return jsonify({'message': '权限更新成功'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': str(e)}), 500

# ------------------- 数据接口 -------------------
@app.route('/classes')
def get_classes():
    classes = [d for d in os.listdir(app.config['UPLOAD_FOLDER']) 
              if os.path.isdir(os.path.join(app.config['UPLOAD_FOLDER'], d))]
    return jsonify({'classes': classes})

@app.route('/subjects')
def get_subjects():
    class_name = request.args.get('class')
    class_dir = os.path.join(app.config['UPLOAD_FOLDER'], class_name)
    subjects = [f[:-5] for f in os.listdir(class_dir) if f.endswith('.xlsx')]
    return jsonify({'subjects': subjects})

@app.route('/students', methods=['GET'])
def get_students():
    """获取学生数据"""
    class_name = request.args.get('class')
    subject = request.args.get('subject')
    assignment = request.args.get('assignment')
    
    try:
        file_path = get_file_path(class_name, subject)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 查找作业列
        col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == assignment:
                col_index = idx
                break
        
        students = []
        for row in ws.iter_rows(min_row=2):
            students.append({
                'id': row[0].value,
                'name': row[1].value,
                'gender': row[2].value,
                'status': row[col_index-1].value if col_index else None
            })
        
        return jsonify({'students': students})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/assignments', methods=['GET'])
def get_assignments():
    """获取作业列表"""
    class_name = request.args.get('class')
    subject = request.args.get('subject')
    
    try:
        file_path = get_file_path(class_name, subject)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 从第四列开始获取作业（前三列：学号、姓名、性别）
        assignments = [cell.value for cell in ws[1][3:]]
        return jsonify({'assignments': assignments})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download', methods=['GET'])
def download_file():
    """文件下载（添加参数验证）"""
    class_name = request.args.get('class')
    subject = request.args.get('subject')
    
    if not class_name or not subject:
        return jsonify({'error': '缺少班级或学科参数'}), 400
    
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], class_name, f"{subject}.xlsx")
        return send_from_directory(
            directory=os.path.dirname(file_path),
            path=os.path.basename(file_path),
            as_attachment=True
        )
    except FileNotFoundError:
        return jsonify({'error': '文件不存在'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/create_assignment', methods=['POST'])
def create_assignment():
    """创建新作业列"""
    data = request.json
    class_name = data['class']
    subject = data['subject']
    assignment = data['assignment']
    
    try:
        file_path = get_file_path(class_name, subject)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 检查重复
        existing = [cell.value for cell in ws[1][3:]]
        if assignment in existing:
            return jsonify({'error': '作业已存在'}), 400
        
        # 找到新列位置（从D列开始）
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value=assignment)
        
        # 初始化空值
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=new_col, value="")
        
        wb.save(file_path)
        return jsonify({'message': '创建成功', 'column': new_col})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update', methods=['POST'])
def update_status():
    """更新学生状态"""
    data = request.json
    class_name = data['class']
    subject = data['subject']
    assignment = data['assignment']
    student_id = data['student_id']
    status = data['status']
    
    try:
        file_path = get_file_path(class_name, subject)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 查找作业列
        col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == assignment:
                col_index = idx
                break
        
        # 查找学生行
        row_index = None
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(student_id):
                row_index = idx
                break
        
        if col_index and row_index:
            ws.cell(row=row_index, column=col_index, value=status)
            wb.save(file_path)
            return jsonify({'message': '更新成功'})
        
        return jsonify({'error': '未找到数据'}), 404
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

#-----------

@app.route('/users', methods=['GET'])
def get_users():
    users = User.query.all()
    return jsonify({
        'users': [{
            'username': u.username,
            'roles': u.roles.split(','),
            'permissions': u.permissions.split(',') if u.permissions else []
        } for u in users]
    }), 200

@app.route('/change_password', methods=['POST'])
def change_password():
    """修改密码接口"""
    data = request.json
    required = ['username', 'new_password', 'operator']
    
    # 检查必要参数
    if not all(k in data for k in required):
        return jsonify({'message': '缺少必要参数'}), 400

    # 验证操作者权限
    operator = User.query.filter_by(username=data['operator']).first()
    if not operator or 'admin' not in operator.roles.split(','):
        return jsonify({'message': '权限不足'}), 403

    # 查找目标用户
    user = User.query.filter_by(username=data['username']).first()
    if not user:
        return jsonify({'message': '用户不存在'}), 404

    try:
        user.password = data['new_password']
        db.session.commit()
        return jsonify({'message': '密码修改成功'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    """文件上传接口"""
    # 验证权限
    operator = request.form.get('operator')
    if not User.query.filter_by(username=operator, roles='admin').first():
        return jsonify({'message': '无上传权限'}), 403

    # 检查文件
    if 'file' not in request.files:
        return jsonify({'message': '未选择文件'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': '无效文件名'}), 400

    # 参数验证
    class_name = request.form.get('class')
    subject = request.form.get('subject')
    if not class_name or not subject:
        return jsonify({'message': '缺少班级或学科参数'}), 400

    # 安全处理路径
    safe_class = secure_filename(class_name).replace('..', '')
    safe_subject = secure_filename(subject).replace('..', '')
    
    # 创建保存目录
    save_dir = os.path.join(app.config['UPLOAD_FOLDER'], safe_class)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    # 保存文件
    filename = f"{safe_subject}.xlsx"
    save_path = os.path.join(save_dir, filename)
    try:
        file.save(save_path)
        return jsonify({'message': '文件上传成功', 'path': save_path}), 200
    except Exception as e:
        return jsonify({'message': f'保存失败: {str(e)}'}), 500

# ------------------- 文件操作接口 -------------------
# 服务器端新增接口
# ----------------- 网站接口 -----------------

@app.route('/zgzxhomeworkmgr')
def homework_mgr():
    """返回中国中学作业管理系统的展示页面"""
    return render_template_string('''
    <!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>龙解软控-中国中学作业管理系统</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: 'Microsoft YaHei', sans-serif;
        }
        
        body {
            background-color: #f5f7fa;
            color: #333;
            line-height: 1.6;
        }
        
        .container {
            width: 100%;
            max-width: 1200px;
            margin: 0 auto;
            padding: 0 20px;
        }
        
        /* 导航栏样式 */
        header {
            background: linear-gradient(135deg, #1a56db 0%, #0e2a5e 100%);
            color: white;
            padding: 1rem 0;
            position: sticky;
            top: 0;
            z-index: 100;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
        }
        
        .navbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .logo {
            display: flex;
            align-items: center;
            gap: 10px;
            font-size: 1.8rem;
            font-weight: 700;
        }
        
        .logo i {
            color: #ff6b35;
        }
        
        .nav-links {
            display: flex;
            gap: 2rem;
        }
        
        .nav-links a {
            color: white;
            text-decoration: none;
            font-weight: 500;
            transition: all 0.3s ease;
            padding: 0.5rem 1rem;
            border-radius: 4px;
        }
        
        .nav-links a:hover {
            background-color: rgba(255, 255, 255, 0.1);
        }
        
        /* 标题区域 */
        .hero {
            background: linear-gradient(rgba(10, 25, 47, 0.8), rgba(10, 25, 47, 0.9));
            color: white;
            padding: 4rem 0;
            text-align: center;
        }
        
        .hero h1 {
            font-size: 2.8rem;
            margin-bottom: 1.5rem;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .hero p {
            font-size: 1.2rem;
            max-width: 800px;
            margin: 0 auto;
            opacity: 0.9;
        }
        
        /* 封面展示区域 */
        .covers {
            padding: 4rem 0;
            background-color: white;
        }
        
        .section-title {
            text-align: center;
            margin-bottom: 3rem;
        }
        
        .section-title h2 {
            font-size: 2.2rem;
            color: #0e2a5e;
            margin-bottom: 1rem;
        }
        
        .section-title p {
            color: #6b7280;
            max-width: 700px;
            margin: 0 auto;
        }
        
        .covers-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 2rem;
        }
        
        .cover-card {
            background: white;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 5px 15px rgba(0, 0, 0, 0.05);
            transition: all 0.3s ease;
        }
        
        .cover-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.1);
        }
        
        .cover-image {
            height: 400px;
            background-color: #f0f4f8;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        
        .cover-image img {
            width: 100%;
            height: 100%;
            object-fit: contain;
        }
        
        .cover-content {
            padding: 1.5rem;
            text-align: center;
        }
        
        .cover-content h3 {
            font-size: 1.3rem;
            margin-bottom: 0.5rem;
            color: #0e2a5e;
        }
        
        /* 界面展示区域 */
        .interfaces {
            padding: 4rem 0;
            background-color: #f0f4f8;
        }
        
        .interface-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 2rem;
        }
        
        .interface-card {
            background: white;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 5px 15px rgba(0, 0, 0, 0.05);
        }
        
        .interface-image {
            height: 500px;
            background-color: #f0f4f8;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        
        .interface-image img {
            width: 100%;
            height: 100%;
            object-fit: contain;
        }
        
        .interface-content {
            padding: 1.5rem;
        }
        
        .interface-content h3 {
            font-size: 1.3rem;
            margin-bottom: 1rem;
            color: #0e2a5e;
        }
        
        /* 页脚样式 */
        footer {
            background-color: #1e293b;
            color: white;
            padding: 3rem 0 1rem;
        }
        
        .footer-content {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 2rem;
            margin-bottom: 2rem;
        }
        
        .footer-column h3 {
            font-size: 1.3rem;
            margin-bottom: 1.5rem;
            position: relative;
        }
        
        .footer-column h3::after {
            content: '';
            position: absolute;
            bottom: -8px;
            left: 0;
            width: 50px;
            height: 3px;
            background-color: #ff6b35;
        }
        
        .footer-column p {
            margin-bottom: 1rem;
            opacity: 0.8;
        }
        
        .copyright {
            text-align: center;
            padding-top: 2rem;
            border-top: 1px solid rgba(255, 255, 255, 0.1);
            opacity: 0.7;
        }
        
        /* 响应式设计 */
        @media (max-width: 768px) {
            .nav-links {
                display: none;
            }
            
            .hero h1 {
                font-size: 2.2rem;
            }
            
            .interface-grid {
                grid-template-columns: 1fr;
            }
        }
    </style>
</head>
<body>
    <!-- 导航栏 -->
    <header>
        <div class="container">
            <nav class="navbar">
                <div class="logo">
                    <span>龙解软控-中国中学作业管理系统</span>
                </div>
                <div class="nav-links">
                    <a href="#home">首页</a>
                    <a href="#covers">软件封面</a>
                    <a href="#interfaces">软件界面</a>
                </div>
            </nav>
        </div>
    </header>

    <!-- 标题区域 -->
    <section class="hero" id="home">
        <div class="container">
            <h1>中国中学作业管理系统</h1>
            <p>为中国中学学生设计的智能作业管理解决方案</p>
        </div>
    </section>

    <!-- 封面展示区域 -->
    <section class="covers" id="covers">
        <div class="container">
            <div class="section-title">
                <h2>软件版本封面</h2>
                <p>不同版本的软件封面设计展示</p>
            </div>
            <div class="covers-grid">
                <div class="cover-card">
                    <div class="cover-image">
                        <img src="static/images/ico1.png" alt="基础版封面">
                    </div>
                    <div class="cover-content">
                        <h3>管理端</h3>
                        <p>快捷管理所有账号及权限</p>
                    </div>
                </div>
                <div class="cover-card">
                    <div class="cover-image">
                        <img src="static/images/ico2.png" alt="专业版封面">
                    </div>
                    <div class="cover-content">
                        <h3>课代表教师端</h3>
                        <p>简单操作管理作业情况</p>
                    </div>
                </div>
                <div class="cover-card">
                    <div class="cover-image">
                        <img src="static/images/ico3.png" alt="旗舰版封面">
                    </div>
                    <div class="cover-content">
                        <h3>学生端</h3>
                        <p>快速查看作业情况</p>
                    </div>
                </div>
            </div>
        </div>
    </section>

    <!-- 界面展示区域 -->
    <section class="interfaces" id="interfaces">
        <div class="container">
            <div class="section-title">
                <h2>软件界面展示</h2>
                <p>系统核心功能界面预览</p>
            </div>
            <div class="interface-grid">
                <div class="interface-card">
                    <div class="interface-image">
                        <img src="static/images/素材1.jpg" alt="作业布置界面">
                    </div>
                    <div class="interface-content">
                        <h3>课代表教师端（电脑）界面</h3>
                        <p>课代表、教师快速创建编辑作业</p>
                    </div>
                </div>
                <div class="interface-card">
                    <div class="interface-image">
                        <img src="static/images/素材2.jpg" alt="学生端（手机）界面">
                    </div>
                    <div class="interface-content">
                        <h3>学生端（手机）界面</h3>
                        <p>学生在线提交作业，教师高效批改与反馈</p>
                    </div>
                </div>
            </div>
        </div>
    </section>

    <!-- 页脚 -->
    <footer>
        <div class="container">
            <div class="footer-content">
                <div class="footer-column">
                    <h3>中国中学作业管理系统</h3>
                    <p>为中国中学学生设计的智能作业管理平台</p>
                </div>
                <div class="footer-column">
                    <h3>联系我们</h3>
                    <p>电话: 17717913079</p>
                    <p>邮箱: luiyixu@163.com</p>
                    <p>地址: 上海市中国中学</p>
                </div>
            </div>
            <div class="copyright">
                <p>&copy; 2025 龙解软控 版权所有</p>
            </div>
        </div>
    </footer>

    <script>
        // 平滑滚动
        document.querySelectorAll('a[href^="#"]').forEach(anchor => {
            anchor.addEventListener('click', function (e) {
                e.preventDefault();
                document.querySelector(this.getAttribute('href')).scrollIntoView({
                    behavior: 'smooth'
                });
            });
        });
    </script>
</body>
</html>

    ''')
# ----------------- 文件下载 -----------------
# ----------------- 初始化服务 -----------------
if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    with app.app_context():
        db.create_all()
    serve(app, host='0.0.0.0', port=5000)
