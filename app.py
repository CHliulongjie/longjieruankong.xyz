from flask import Flask, request, jsonify, send_from_directory, render_template 
from flask_sqlalchemy import SQLAlchemy
import sqlite3
import os
import logging
from waitress import serve
import openpyxl
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

# ===== 路径配置 =====
base_dir = os.path.dirname(os.path.abspath(__file__))
data_dir = os.path.join(base_dir, 'data')
accounts_dir = os.path.join(base_dir, 'accounts')
permissions_dir = os.path.join(base_dir, 'permissions')
os.makedirs(data_dir, exist_ok=True)
os.makedirs(accounts_dir, exist_ok=True)
os.makedirs(permissions_dir, exist_ok=True)

# ===== 数据库配置 =====
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(data_dir, "users.db").replace("\\", "/")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'data'
ALLOWED_EXTENSIONS = {'xlsx'}
db = SQLAlchemy(app)

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    roles = db.Column(db.String(200))
    permissions = db.Column(db.String(200))

# ===== 工具函数 =====
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    return conn

def get_file_path(class_name, subject):
    return os.path.join(app.config['UPLOAD_FOLDER'], class_name, f"{subject}.xlsx")

# ------------------- 核心接口 -------------------
@app.route('/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    app_name = data.get('app_name', 'homework')
    
    admin_db_path = os.path.join(accounts_dir, 'admin.db')
    if os.path.exists(admin_db_path):
        admin_conn = sqlite3.connect(admin_db_path)
        admin_cursor = admin_conn.cursor()
        admin_cursor.execute("SELECT uid, password, admin_level FROM admins WHERE username = ?", (username,))
        admin_result = admin_cursor.fetchone()
        admin_conn.close()
        
        if admin_result:
            uid, stored_password, admin_level = admin_result
            if password == stored_password:
                perm_db_path = os.path.join(permissions_dir, f'{app_name}_permissions.db')
                if os.path.exists(perm_db_path):
                    perm_conn = sqlite3.connect(perm_db_path)
                    perm_cursor = perm_conn.cursor()
                    perm_cursor.execute(f"SELECT permissions FROM {app_name}_permissions WHERE uid = ?", (uid,))
                    perm_result = perm_cursor.fetchone()
                    
                    if perm_result:
                        permissions = perm_result[0]
                        is_new = False
                    else:
                        permissions = '*' if admin_level == 'super_admin' else '.'
                        perm_cursor.execute(f"INSERT INTO {app_name}_permissions (uid, permissions) VALUES (?, ?)", (uid, permissions))
                        perm_conn.commit()
                        is_new = True
                    perm_conn.close()
                    
                    return jsonify({
                        'success': True,
                        'username': username,
                        'uid': uid,
                        'account_type': 'admin',
                        'admin_level': admin_level,
                        'permissions': permissions,
                        'message': '登录成功',
                        'app_name': app_name,
                        'is_new_permission': is_new
                    }), 200
                else:
                    return jsonify({'success': False, 'message': '权限数据库不存在'}), 500
    
    user_db_path = os.path.join(accounts_dir, 'users.db')
    if os.path.exists(user_db_path):
        user_conn = sqlite3.connect(user_db_path)
        user_cursor = user_conn.cursor()
        user_cursor.execute("SELECT uid, password FROM users WHERE username = ?", (username,))
        user_result = user_cursor.fetchone()
        user_conn.close()
        
        if user_result:
            uid, stored_password = user_result
            if password == stored_password:
                perm_db_path = os.path.join(permissions_dir, f'{app_name}_permissions.db')
                if os.path.exists(perm_db_path):
                    perm_conn = sqlite3.connect(perm_db_path)
                    perm_cursor = perm_conn.cursor()
                    perm_cursor.execute(f"SELECT permissions FROM {app_name}_permissions WHERE uid = ?", (uid,))
                    perm_result = perm_cursor.fetchone()
                    
                    if perm_result:
                        permissions = perm_result[0]
                        is_new = False
                    else:
                        permissions = '.'
                        perm_cursor.execute(f"INSERT INTO {app_name}_permissions (uid, permissions) VALUES (?, ?)", (uid, permissions))
                        perm_conn.commit()
                        is_new = True
                    perm_conn.close()
                    
                    return jsonify({
                        'success': True,
                        'username': username,
                        'uid': uid,
                        'account_type': 'user',
                        'permissions': permissions,
                        'message': '登录成功',
                        'app_name': app_name,
                        'is_new_permission': is_new
                    }), 200
                else:
                    return jsonify({'success': False, 'message': '权限数据库不存在'}), 500
    
    return jsonify({'success': False, 'message': '用户名或密码错误'}), 401

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    account_type = data.get('account_type', 'user')
    
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400
    
    admin_db_path = os.path.join(accounts_dir, 'admin.db')
    user_db_path = os.path.join(accounts_dir, 'users.db')
    
    if os.path.exists(admin_db_path):
        admin_conn = sqlite3.connect(admin_db_path)
        admin_cursor = admin_conn.cursor()
        admin_cursor.execute("SELECT username FROM admins WHERE username = ?", (username,))
        if admin_cursor.fetchone():
            admin_conn.close()
            return jsonify({'success': False, 'message': '用户名已存在'}), 400
    
    if os.path.exists(user_db_path):
        user_conn = sqlite3.connect(user_db_path)
        user_cursor = user_conn.cursor()
        user_cursor.execute("SELECT username FROM users WHERE username = ?", (username,))
        if user_cursor.fetchone():
            user_conn.close()
            if 'admin_conn' in locals():
                admin_conn.close()
            return jsonify({'success': False, 'message': '用户名已存在'}), 400
    
    from datetime import datetime
    now = datetime.now()
    date_str = now.strftime('%y%m%d')
    
    if account_type == 'admin':
        admin_conn = sqlite3.connect(admin_db_path)
        admin_cursor = admin_conn.cursor()
        admin_cursor.execute("SELECT COUNT(*) FROM admins")
        count = admin_cursor.fetchone()[0]
        uid = f'LJRK-SFQD-{date_str}{count + 1:04d}'
        
        admin_cursor.execute(
            "INSERT INTO admins (username, password, uid, admin_level, register_date) VALUES (?, ?, ?, ?, ?)",
            (username, password, uid, 'admin', now.strftime('%Y-%m-%d'))
        )
        admin_conn.commit()
        admin_conn.close()
        
        perm_db_path = os.path.join(permissions_dir, 'homework_permissions.db')
        if os.path.exists(perm_db_path):
            perm_conn = sqlite3.connect(perm_db_path)
            perm_cursor = perm_conn.cursor()
            perm_cursor.execute("INSERT INTO homework_permissions (uid, permissions) VALUES (?, ?)", (uid, '.'))
            perm_conn.commit()
            perm_conn.close()
        
        return jsonify({
            'success': True,
            'message': '管理员账户创建成功',
            'uid': uid,
            'account_type': 'admin'
        }), 201
    
    else:
        user_conn = sqlite3.connect(user_db_path)
        user_cursor = user_conn.cursor()
        user_cursor.execute("SELECT COUNT(*) FROM users")
        count = user_cursor.fetchone()[0]
        uid = f'LJRK-SFQD-{date_str}{count + 1:04d}'
        
        user_cursor.execute(
            "INSERT INTO users (username, password, uid, register_date) VALUES (?, ?, ?, ?)",
            (username, password, uid, now.strftime('%Y-%m-%d'))
        )
        user_conn.commit()
        user_conn.close()
        
        perm_db_path = os.path.join(permissions_dir, 'homework_permissions.db')
        if os.path.exists(perm_db_path):
            perm_conn = sqlite3.connect(perm_db_path)
            perm_cursor = perm_conn.cursor()
            perm_cursor.execute("INSERT INTO homework_permissions (uid, permissions) VALUES (?, ?)", (uid, '.'))
            perm_conn.commit()
            perm_conn.close()
        
        return jsonify({
            'success': True,
            'message': '用户账户创建成功',
            'uid': uid,
            'account_type': 'user'
        }), 201

@app.route('/set_permissions', methods=['POST'])
def set_permissions():
    data = request.json
    uid = data.get('uid')
    permissions = data.get('permissions')
    app_name = data.get('app_name', 'homework')
    
    if not uid:
        return jsonify({'success': False, 'message': '用户识别码不能为空'}), 400
    
    perm_db_path = os.path.join(permissions_dir, f'{app_name}_permissions.db')
    if os.path.exists(perm_db_path):
        perm_conn = sqlite3.connect(perm_db_path)
        perm_cursor = perm_conn.cursor()
        
        try:
            perm_cursor.execute(f"""
                INSERT INTO {app_name}_permissions (uid, permissions, updated_at)
                VALUES (?, ?, datetime('now'))
                ON CONFLICT(uid) DO UPDATE SET
                    permissions = excluded.permissions,
                    updated_at = datetime('now')
            """, (uid, permissions))
            perm_conn.commit()
            perm_conn.close()
            return jsonify({'success': True, 'message': '权限更新成功'}), 200
        except Exception as e:
            perm_conn.close()
            return jsonify({'success': False, 'message': str(e)}), 500
    else:
        return jsonify({'success': False, 'message': '权限数据库不存在'}), 500

@app.route('/get_permissions', methods=['POST'])
def get_permissions():
    data = request.json
    uid = data.get('uid')
    app_name = data.get('app_name', 'homework')
    
    if not uid:
        return jsonify({'success': False, 'message': '用户识别码不能为空'}), 400
    
    perm_db_path = os.path.join(permissions_dir, f'{app_name}_permissions.db')
    if os.path.exists(perm_db_path):
        perm_conn = sqlite3.connect(perm_db_path)
        perm_cursor = perm_conn.cursor()
        perm_cursor.execute(f"SELECT permissions FROM {app_name}_permissions WHERE uid = ?", (uid,))
        result = perm_cursor.fetchone()
        perm_conn.close()
        
        if result:
            return jsonify({'success': True, 'permissions': result[0]}), 200
        else:
            return jsonify({'success': False, 'message': '未找到权限记录'}), 404
    else:
        return jsonify({'success': False, 'message': '权限数据库不存在'}), 500

@app.route('/change_password', methods=['POST'])
def change_password():
    data = request.json
    uid = data.get('uid')
    new_password = data.get('new_password')
    
    if not uid or not new_password:
        return jsonify({'success': False, 'message': '识别码和新密码不能为空'}), 400
    
    admin_db_path = os.path.join(accounts_dir, 'admin.db')
    if os.path.exists(admin_db_path):
        admin_conn = sqlite3.connect(admin_db_path)
        admin_cursor = admin_conn.cursor()
        admin_cursor.execute("UPDATE admins SET password = ? WHERE uid = ?", (new_password, uid))
        
        if admin_cursor.rowcount > 0:
            admin_conn.commit()
            admin_conn.close()
            return jsonify({'success': True, 'message': '密码修改成功'}), 200
        admin_conn.close()
    
    user_db_path = os.path.join(accounts_dir, 'users.db')
    if os.path.exists(user_db_path):
        user_conn = sqlite3.connect(user_db_path)
        user_cursor = user_conn.cursor()
        user_cursor.execute("UPDATE users SET password = ? WHERE uid = ?", (new_password, uid))
        
        if user_cursor.rowcount > 0:
            user_conn.commit()
            user_conn.close()
            return jsonify({'success': True, 'message': '密码修改成功'}), 200
        user_conn.close()
    
    return jsonify({'success': False, 'message': '用户不存在'}), 404

@app.route('/users', methods=['GET'])
def get_users():
    users_data = []
    app_name = request.args.get('app_name', 'homework')
    
    # 获取权限数据库路径
    perm_db_path = os.path.join(permissions_dir, f'{app_name}_permissions.db')
    
    def get_user_permissions(uid):
        """从权限数据库获取用户权限"""
        if os.path.exists(perm_db_path):
            try:
                perm_conn = sqlite3.connect(perm_db_path)
                perm_cursor = perm_conn.cursor()
                perm_cursor.execute(f"SELECT permissions FROM {app_name}_permissions WHERE uid = ?", (uid,))
                result = perm_cursor.fetchone()
                perm_conn.close()
                if result and result[0]:
                    # 将逗号分隔的字符串转换为数组
                    return result[0].split(',')
                return ['.']
            except:
                return ['.']
        return ['.']
    
    admin_db_path = os.path.join(accounts_dir, 'admin.db')
    if os.path.exists(admin_db_path):
        admin_conn = sqlite3.connect(admin_db_path)
        admin_cursor = admin_conn.cursor()
        admin_cursor.execute("SELECT username, uid, admin_level, register_date FROM admins")
        for row in admin_cursor.fetchall():
            permissions = get_user_permissions(row[1])
            users_data.append({
                'username': row[0],
                'uid': row[1],
                'roles': ['admin'],
                'permissions': permissions,
                'account_type': 'admin',
                'admin_level': row[2]
            })
        admin_conn.close()
    
    user_db_path = os.path.join(accounts_dir, 'users.db')
    if os.path.exists(user_db_path):
        user_conn = sqlite3.connect(user_db_path)
        user_cursor = user_conn.cursor()
        user_cursor.execute("SELECT username, uid, register_date FROM users")
        for row in user_cursor.fetchall():
            permissions = get_user_permissions(row[1])
            users_data.append({
                'username': row[0],
                'uid': row[1],
                'roles': ['user'],
                'permissions': permissions,
                'account_type': 'user'
            })
        user_conn.close()
    
    return jsonify({'users': users_data}), 200

@app.route('/classes')
def get_classes():
    classes = []
    for item in os.listdir(app.config['UPLOAD_FOLDER']):
        item_path = os.path.join(app.config['UPLOAD_FOLDER'], item)
        if os.path.isdir(item_path):
            classes.append(item)
    return jsonify({'classes': sorted(classes)})

@app.route('/subjects')
def get_subjects():
    class_name = request.args.get('class_name')
    
    # 如果提供了班级名称，从对应的班级文件夹中获取学科
    if class_name:
        class_dir = os.path.join(app.config['UPLOAD_FOLDER'], class_name)
        subjects = []
        if os.path.exists(class_dir):
            for file in os.listdir(class_dir):
                if file.endswith('.xlsx'):
                    # 去掉 .xlsx 后缀得到学科名称
                    subject = file.replace('.xlsx', '')
                    subjects.append(subject)
        return jsonify({'subjects': sorted(subjects)})
    
    # 如果没有提供班级名称，返回所有班级中存在的学科（去重）
    all_subjects = set()
    for item in os.listdir(app.config['UPLOAD_FOLDER']):
        item_path = os.path.join(app.config['UPLOAD_FOLDER'], item)
        if os.path.isdir(item_path):
            for file in os.listdir(item_path):
                if file.endswith('.xlsx'):
                    subject = file.replace('.xlsx', '')
                    all_subjects.add(subject)
    return jsonify({'subjects': sorted(list(all_subjects))})

@app.route('/students', methods=['GET'])
def get_students():
    class_name = request.args.get('class_name')
    if not class_name:
        return jsonify({'error': '缺少班级名称参数'}), 400
    
    students = []
    class_dir = os.path.join(app.config['UPLOAD_FOLDER'], class_name)
    
    if not os.path.exists(class_dir):
        return jsonify({'students': students})
    
    for file in os.listdir(class_dir):
        if file.endswith('.xlsx'):
            file_path = os.path.join(class_dir, file)
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        student = {'name': row[0], 'id': str(row[1]) if row[1] else ''}
                        if student not in students:
                            students.append(student)
                workbook.close()
            except Exception as e:
                logging.error(f"读取Excel文件失败: {e}")
    
    students = sorted(students, key=lambda x: x['name'])
    return jsonify({'students': students})

@app.route('/assignments', methods=['GET'])
def get_assignments():
    class_name = request.args.get('class_name')
    subject = request.args.get('subject')
    
    if not class_name or not subject:
        return jsonify({'error': '缺少班级或学科参数'}), 400
    
    file_path = get_file_path(class_name, subject)
    if not os.path.exists(file_path):
        return jsonify({'error': '文件不存在'}), 404
    
    assignments = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # 只获取作业名称列表（从第4列开始，即第1行的表头）
        for col in range(4, sheet.max_column + 1):
            assignment_name = sheet.cell(row=1, column=col).value
            if assignment_name:
                assignments.append(assignment_name)
        
        workbook.close()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
    return jsonify({'assignments': assignments})

@app.route('/download', methods=['GET'])
def download_file():
    class_name = request.args.get('class_name')
    subject = request.args.get('subject')
    
    if not class_name or not subject:
        return jsonify({'error': '缺少班级或学科参数'}), 400
    
    file_path = get_file_path(class_name, subject)
    if not os.path.exists(file_path):
        return jsonify({'error': '文件不存在'}), 404
    
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    
    return send_from_directory(directory, filename, as_attachment=True)

@app.route('/create_assignment', methods=['POST'])
def create_assignment():
    data = request.json
    class_name = data.get('class_name')
    subject = data.get('subject')
    assignments = data.get('assignments', [])
    
    if not class_name or not subject:
        return jsonify({'error': '缺少班级或学科参数'}), 400
    
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], class_name), exist_ok=True)
    file_path = get_file_path(class_name, subject)
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        if assignments:
            headers = list(assignments[0].keys())
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            for row_idx, assignment in enumerate(assignments, 2):
                for col_idx, header in enumerate(headers, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=assignment.get(header))
        
        workbook.save(file_path)
        workbook.close()
        
        return jsonify({'message': '作业文件创建成功'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update', methods=['POST'])
def update_assignment():
    data = request.json
    class_name = data.get('class_name')
    subject = data.get('subject')
    assignments = data.get('assignments', [])
    
    if not class_name or not subject:
        return jsonify({'error': '缺少班级或学科参数'}), 400
    
    file_path = get_file_path(class_name, subject)
    if not os.path.exists(file_path):
        return jsonify({'error': '文件不存在'}), 404
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        if assignments:
            headers = list(assignments[0].keys())
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            for row_idx, assignment in enumerate(assignments, 2):
                for col_idx, header in enumerate(headers, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=assignment.get(header))
        
        workbook.save(file_path)
        workbook.close()
        
        return jsonify({'message': '作业文件更新成功'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    class_name = request.form.get('class_name')
    subject = request.form.get('subject')
    
    if not class_name or not subject:
        return jsonify({'error': '缺少班级或学科参数'}), 400
    
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名不能为空'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(f"{subject}.xlsx")
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], class_name), exist_ok=True)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], class_name, filename)
        file.save(file_path)
        return jsonify({'message': '文件上传成功'}), 200
    else:
        return jsonify({'error': '不支持的文件格式'}), 400

# ----------------- 初始化服务 -----------------
if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG)
    
    with app.app_context():
        db.create_all()
    
    port = 8000
    print(f"启动作业管理系统服务，端口 {port}...")
    print(f"内部访问地址: http://localhost:{port}")
    print("Nginx将处理HTTPS并代理到此端口")
    
    serve(app, host='127.0.0.1', port=port, threads=4)
